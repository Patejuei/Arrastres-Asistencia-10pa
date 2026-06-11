import datetime
from typing import Any, cast
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.database.db_service import db_service
from src.utils.date_helpers import parse_date

# Estilos de Celda suizos (Rojo Suizo y Gris Oscuro)
HEADER_FILL = PatternFill(start_color="D81E05", end_color="D81E05", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1E1E24")
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="555555")
DATA_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC")
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

def auto_fit_columns(ws, max_padding=4, min_width=10):
    """Ajusta automáticamente el ancho de las columnas en base al contenido."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + max_padding, min_width)

# =========================================================================
# 1. INFORME ARRASTRE DE ASISTENCIA
# =========================================================================
def generate_arrastre_excel(file_path: str, fecha_inicio: str, fecha_fin: str) -> tuple[bool, str]:
    """
    Genera el Excel del Arrastre de Asistencia en un periodo.
    Columnas: Bombero, Universo de Efectivas, Total de listas efectivas, Total de Listas de abono.
    """
    try:
        f_ini = parse_date(fecha_inicio)
        f_fin = parse_date(fecha_fin)
        
        # 1. Obtener todos los bomberos
        bomberos: list[dict[str, Any]] = db_service.get_bomberos()
        
        # 2. Obtener todos los actos de servicio de tipo 'Efectiva' y 'Abono' en el rango de fechas
        # Para evitar sobrecargar la BD, traemos todos los actos en el rango
        # Traemos todos los actos en la BD y filtramos localmente (ya que postgrest select eq etc ya lo hace en db_service)
        # Vamos a obtener los actos filtrando por rango directamente.
        # Implementamos una consulta directa de actos en el rango.
        supabase = db_service._get_supabase()
        res_actos = supabase.table("actos_servicio").select("*").gte("fecha", fecha_inicio).lte("fecha", fecha_fin).execute()
        actos: list[dict[str, Any]] = cast(list[dict[str, Any]], res_actos.data or [])
        
        # 3. Obtener todas las asistencias de esos actos
        asistencias: list[dict[str, Any]] = []
        if actos:
            actos_ids = [a["id"] for a in actos]
            # Consultas en Supabase para obtener las asistencias
            res_asis = supabase.table("asistencias").select("*").in_("id_acto", actos_ids).execute()
            asistencias = cast(list[dict[str, Any]], res_asis.data or [])
        else:
            asistencias = []
            
        # Mapear asistencias por (id_acto, reg_bombero) -> cantidad_listas
        asis_map = {}
        for asis in asistencias:
            asis_map[(asis["id_acto"], asis["reg_bombero"])] = asis["cantidad_listas"]
            
        # 4. Obtener todos los movimientos de personal para calcular universo de efectivas
        res_movs = supabase.table("movimientos_personal").select("*").execute()
        movs: list[dict[str, Any]] = cast(list[dict[str, Any]], res_movs.data or [])
        
        # Organizar movimientos por bombero
        movs_por_bombero: dict[str, list[dict[str, Any]]] = {}
        for m in movs:
            reg = m["reg_bombero"]
            if isinstance(reg, str):
                if reg not in movs_por_bombero:
                    movs_por_bombero[reg] = []
                movs_por_bombero[reg].append(m)

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Arrastre Asistencia"
        
        # Títulos
        ws.cell(row=1, column=1, value='Décima Compañía "Bomba Suiza"').font = TITLE_FONT
        ws.cell(row=2, column=1, value=f"Arrastre de Asistencia: {f_ini.strftime('%d-%m-%Y')} al {f_fin.strftime('%d-%m-%Y')}").font = SUBTITLE_FONT
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 18
        
        # Encabezados
        headers = ["Registro General", "Bombero", "Universo de Efectivas", "Total Listas Efectivas", "Total Listas Abono", "Asistencia Neta (%)"]
        ws.row_dimensions[4].height = 24
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
            
        row_idx = 5
        for b in bomberos:
            reg = b["registro_general"]
            nombre_bombero = f"{b['apellido_paterno']} {b['apellido_materno']}, {b['nombres']}"
            estado = b.get("estado", "Activo")
            
            # Solo procesar bomberos que no estén en estado "Renunciado" o similar en la fecha del reporte,
            # pero mejor procesar a todos y calcular su universo
            # Universo de listas efectivas: contar actos de tipo "Efectiva" que ocurrieron mientras el bombero estuvo activo
            universo_efectivas = 0
            listas_efectivas = 0
            listas_abono = 0
            
            # Obtener periodos activos del bombero
            periodos_activos = []
            b_movs = movs_por_bombero.get(reg, [])
            for m in b_movs:
                ingreso = parse_date(m["fecha_ingreso"])
                salida = parse_date(m["fecha_salida"]) if m["fecha_salida"] else None
                periodos_activos.append((ingreso, salida))
                
            # Si no tiene movimientos registrados, asumimos que estuvo activo siempre
            if not periodos_activos:
                # O si ingresó antes, simulamos un periodo desde el inicio del universo
                periodos_activos.append((datetime.date(1900, 1, 1), None))
                
            for acto in actos:
                acto_fecha = parse_date(acto["fecha"])
                acto_tipo = acto["tipo"] # 'Efectiva' o 'Abono'
                
                # Verificar si la fecha del acto cae dentro de algún período activo del bombero
                esta_activo_en_acto = False
                for ingreso, salida in periodos_activos:
                    if acto_fecha >= ingreso and (salida is None or acto_fecha <= salida):
                        esta_activo_en_acto = True
                        break
                        
                if esta_activo_en_acto:
                    if acto_tipo == "Efectiva":
                        universo_efectivas += 1
                        
                # Obtener la asistencia real del bombero en este acto
                listas_asis = asis_map.get((acto["id"], reg), 0)
                if listas_asis > 0:
                    if acto_tipo == "Efectiva":
                        listas_efectivas += listas_asis
                    elif acto_tipo == "Abono":
                        listas_abono += listas_asis
            
            # Calcular porcentaje de asistencia
            porcentaje_str = "0.0%"
            if universo_efectivas > 0:
                porcentaje = (listas_efectivas / universo_efectivas) * 100
                porcentaje_str = f"{round(porcentaje, 1)}%"
                
            # Escribir en Excel
            ws.cell(row=row_idx, column=1, value=reg).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=2, value=nombre_bombero).alignment = ALIGN_LEFT
            ws.cell(row=row_idx, column=3, value=universo_efectivas).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=4, value=listas_efectivas).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=5, value=listas_abono).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=6, value=porcentaje_str).alignment = ALIGN_CENTER
            
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).font = DATA_FONT
                ws.cell(row=row_idx, column=c).border = THIN_BORDER
                
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1
            
        auto_fit_columns(ws)
        wb.save(file_path)
        return True, "Reporte de Arrastre de Asistencia generado con éxito."
    except Exception as e:
        return False, f"Error al generar Excel: {str(e)}"

# =========================================================================
# 2. INFORME RESUMEN DE ASISTENCIA POR BOMBERO
# =========================================================================
def generate_resumen_bombero_excel(file_path: str, reg_bombero: str) -> tuple[bool, str]:
    """
    Genera el Resumen de Asistencia de un bombero en particular.
    Hoja 1: Ficha del bombero + conteos de asistencia.
    Hojas 2+: Actos de servicio donde registra asistencia, separados por año del acto.
    """
    try:
        # 1. Obtener la ficha del bombero
        bombero = db_service.get_bombero(reg_bombero)
        if not bombero:
            return False, f"No se encontró al bombero con Registro General '{reg_bombero}'."
            
        # 2. Obtener todas las asistencias del bombero con cantidad_listas > 0
        # (incluyendo los detalles de los actos de servicio asociados)
        supabase = db_service._get_supabase()
        res_asis = supabase.table("asistencias").select("*, actos_servicio(*)").eq("reg_bombero", reg_bombero).gt("cantidad_listas", 0).execute()
        asistencias: list[dict[str, Any]] = cast(list[dict[str, Any]], res_asis.data or [])
        
        # 3. Separar asistencias por año del acto de servicio e ir contando totales
        totales_efectivas = 0
        totales_abonos = 0
        asistencias_por_anio: dict[str, list[dict[str, Any]]] = {}
        
        for asis in asistencias:
            acto = asis.get("actos_servicio")
            if not isinstance(acto, dict):
                continue
                
            fecha_acto = parse_date(acto["fecha"])
            anio = str(fecha_acto.year)
            
            if anio not in asistencias_por_anio:
                asistencias_por_anio[anio] = []
                
            asistencias_por_anio[anio].append({
                "fecha": fecha_acto,
                "corr_general": acto.get("corr_general"),
                "corr_cia": acto.get("corr_cia"),
                "clave": acto.get("clave"),
                "direccion": acto.get("direccion"),
                "esquina": acto.get("esquina"),
                "tipo": acto.get("tipo"),
                "cantidad_listas": asis.get("cantidad_listas", 0)
            })
            
            # Contar listas
            if acto.get("tipo") == "Efectiva":
                totales_efectivas += asis.get("cantidad_listas", 0)
            elif acto.get("tipo") == "Abono":
                totales_abonos += asis.get("cantidad_listas", 0)

        # Ordenar los años de manera descendente
        anios_ordenados = sorted(list(asistencias_por_anio.keys()), reverse=True)

        wb = Workbook()
        
        # Hoja 1: Resumen General / Ficha
        ws_resumen = wb.active
        assert ws_resumen is not None
        ws_resumen.title = "Ficha y Resumen"
        
        # Títulos
        ws_resumen.cell(row=1, column=1, value='Décima Compañía "Bomba Suiza"').font = TITLE_FONT
        ws_resumen.cell(row=2, column=1, value="Ficha Individual de Asistencia").font = SUBTITLE_FONT
        
        # Información del Bombero
        info_labels = [
            ("Registro General:", reg_bombero),
            ("Nombres:", bombero["nombres"]),
            ("Apellidos:", f"{bombero['apellido_paterno']} {bombero['apellido_materno']}"),
            ("RUT:", bombero["rut"]),
            ("Estado Actual:", bombero["estado"]),
            ("Total Listas Efectivas:", totales_efectivas),
            ("Total Listas de Abono:", totales_abonos),
            ("Total Listas Acumuladas:", totales_efectivas + totales_abonos)
        ]
        
        row_idx = 4
        for label, val in info_labels:
            cell_label = ws_resumen.cell(row=row_idx, column=1, value=label)
            cell_label.font = BOLD_FONT
            cell_label.alignment = ALIGN_LEFT
            
            cell_val = ws_resumen.cell(row=row_idx, column=2, value=val)
            cell_val.font = DATA_FONT
            cell_val.alignment = ALIGN_LEFT
            
            # Dar un color especial a los totales
            if "Total" in label:
                cell_label.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                cell_val.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                cell_val.font = BOLD_FONT
                
            cell_label.border = THIN_BORDER
            cell_val.border = THIN_BORDER
            ws_resumen.row_dimensions[row_idx].height = 20
            row_idx += 1
            
        auto_fit_columns(ws_resumen)
        
        # Hojas Siguientes: Una por año
        for anio in anios_ordenados:
            ws_anio = wb.create_sheet(title=f"Año {anio}")
            assert ws_anio is not None
            
            ws_anio.cell(row=1, column=1, value=f"Historial de Actos del Año {anio}").font = TITLE_FONT
            ws_anio.cell(row=2, column=1, value=f"Bombero: {bombero['nombres']} {bombero['apellido_paterno']} (Reg. {reg_bombero})").font = SUBTITLE_FONT
            ws_anio.row_dimensions[1].height = 25
            ws_anio.row_dimensions[2].height = 18
            
            headers = ["Fecha", "Corr. General", "Corr. Compañía", "Clave", "Dirección", "Esquina", "Tipo Acto", "Listas Asistidas"]
            ws_anio.row_dimensions[4].height = 24
            
            for col_idx, text in enumerate(headers, 1):
                cell = ws_anio.cell(row=4, column=col_idx, value=text)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = ALIGN_CENTER
                cell.border = THIN_BORDER
                
            # Ordenar actos del año por fecha ascendente
            actos_anio = sorted(asistencias_por_anio[anio], key=lambda x: x["fecha"])
            
            r_idx = 5
            for acto in actos_anio:
                ws_anio.cell(row=r_idx, column=1, value=acto["fecha"].strftime("%d-%m-%Y")).alignment = ALIGN_CENTER
                ws_anio.cell(row=r_idx, column=2, value=acto["corr_general"] or "-").alignment = ALIGN_CENTER
                ws_anio.cell(row=r_idx, column=3, value=acto["corr_cia"]).alignment = ALIGN_CENTER
                ws_anio.cell(row=r_idx, column=4, value=acto["clave"]).alignment = ALIGN_CENTER
                ws_anio.cell(row=r_idx, column=5, value=acto["direccion"]).alignment = ALIGN_LEFT
                ws_anio.cell(row=r_idx, column=6, value=acto["esquina"] or "-").alignment = ALIGN_LEFT
                ws_anio.cell(row=r_idx, column=7, value=acto["tipo"]).alignment = ALIGN_CENTER
                ws_anio.cell(row=r_idx, column=8, value=acto["cantidad_listas"]).alignment = ALIGN_CENTER
                
                for c in range(1, 9):
                    ws_anio.cell(row=r_idx, column=c).font = DATA_FONT
                    ws_anio.cell(row=r_idx, column=c).border = THIN_BORDER
                    
                ws_anio.row_dimensions[r_idx].height = 18
                r_idx += 1
                
            # Agregar fila de resumen al final del año
            ws_anio.cell(row=r_idx, column=6, value="TOTAL DEL AÑO").font = BOLD_FONT
            ws_anio.cell(row=r_idx, column=6).alignment = ALIGN_RIGHT
            
            total_listas_anio = sum(a["cantidad_listas"] for a in actos_anio)
            cell_tot = ws_anio.cell(row=r_idx, column=8, value=total_listas_anio)
            cell_tot.font = BOLD_FONT
            cell_tot.alignment = ALIGN_CENTER
            
            ws_anio.cell(row=r_idx, column=6).border = THIN_BORDER
            ws_anio.cell(row=r_idx, column=8).border = THIN_BORDER
            ws_anio.row_dimensions[r_idx].height = 20
            
            auto_fit_columns(ws_anio)
            
        wb.save(file_path)
        return True, "Reporte individual generado con éxito."
    except Exception as e:
        return False, f"Error al generar Excel del bombero: {str(e)}"

# =========================================================================
# 3. INFORME LISTADO DE BOMBEROS
# =========================================================================
def generate_listado_bomberos_excel(file_path: str) -> tuple[bool, str]:
    """Genera un listado de todos los bomberos registrados en Excel."""
    try:
        bomberos: list[dict[str, Any]] = db_service.get_bomberos()
        
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Bomberos Registrados"
        
        ws.cell(row=1, column=1, value='Décima Compañía "Bomba Suiza"').font = TITLE_FONT
        ws.cell(row=2, column=1, value="Nómina General de Bomberos").font = SUBTITLE_FONT
        
        headers = ["Registro General", "Nombres", "Apellido Paterno", "Apellido Materno", "RUT", "Estado Actual"]
        ws.row_dimensions[4].height = 24
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
            
        row_idx = 5
        for b in bomberos:
            ws.cell(row=row_idx, column=1, value=b["registro_general"]).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=2, value=b["nombres"]).alignment = ALIGN_LEFT
            ws.cell(row=row_idx, column=3, value=b["apellido_paterno"]).alignment = ALIGN_LEFT
            ws.cell(row=row_idx, column=4, value=b["apellido_materno"]).alignment = ALIGN_LEFT
            ws.cell(row=row_idx, column=5, value=b["rut"]).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=6, value=b["estado"]).alignment = ALIGN_CENTER
            
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).font = DATA_FONT
                ws.cell(row=row_idx, column=c).border = THIN_BORDER
                
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1
            
        auto_fit_columns(ws)
        wb.save(file_path)
        return True, "Nómina de bomberos generada con éxito."
    except Exception as e:
        return False, f"Error al generar Excel de bomberos: {str(e)}"

# =========================================================================
# 4. INFORME LISTADO DE LICENCIAS
# =========================================================================
def generate_listado_licencias_excel(file_path: str) -> tuple[bool, str]:
    """Genera un listado de todas las licencias registradas en Excel."""
    try:
        licencias: list[dict[str, Any]] = db_service.get_licencias()
        
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Licencias Registradas"
        
        ws.cell(row=1, column=1, value='Décima Compañía "Bomba Suiza"').font = TITLE_FONT
        ws.cell(row=2, column=1, value="Listado Histórico de Licencias").font = SUBTITLE_FONT
        
        headers = ["Reg. Bombero", "Nombre Bombero", "Fecha Desde", "Fecha Hasta", "Motivo", "Aprobada"]
        ws.row_dimensions[4].height = 24
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
            
        row_idx = 5
        for lic in licencias:
            reg = lic["reg_bombero"]
            b_info = lic.get("bomberos")
            if not isinstance(b_info, dict):
                b_info = {}
            nombre_bombero = f"{b_info.get('apellido_paterno', '')} {b_info.get('apellido_materno', '')}, {b_info.get('nombres', '')}"
            
            f_desde = parse_date(lic["fecha_desde"]).strftime("%d-%m-%Y")
            f_hasta = parse_date(lic["fecha_hasta"]).strftime("%d-%m-%Y")
            aprobado_str = "Sí" if lic.get("aprobado") else "No / Pendiente"
            
            ws.cell(row=row_idx, column=1, value=reg).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=2, value=nombre_bombero).alignment = ALIGN_LEFT
            ws.cell(row=row_idx, column=3, value=f_desde).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=4, value=f_hasta).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=5, value=lic["motivo"]).alignment = ALIGN_LEFT
            ws.cell(row=row_idx, column=6, value=aprobado_str).alignment = ALIGN_CENTER
            
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).font = DATA_FONT
                ws.cell(row=row_idx, column=c).border = THIN_BORDER
                
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1
            
        auto_fit_columns(ws)
        wb.save(file_path)
        return True, "Historial de licencias generado con éxito."
    except Exception as e:
        return False, f"Error al generar Excel de licencias: {str(e)}"
